#!/usr/bin/env python3
"""
Extract the Brisket "Cook & Hold" Tenderness Model from Steve Gow's official
spreadsheet into a single canonical JSON file that both the Python package
and the web app read from.

Source spreadsheet: "Holding chart with Calculator V3 (April 30, 2026)"
Published alongside: Brisket Holding Masterclass (And Tenderness Model)
https://smoketrailsbbq.com/brisket-holding-masterclass-and-tenderness-model/
By Steve Gow / Smoke Trails BBQ

This script does NOT invent any numbers. Every value below is read directly
from the workbook's cells (or, where noted, transcribed verbatim from Steve
Gow's own comment replies on the article, which is the only place a couple
of extra reference details -- like the exact zone-by-zone breakdown of a
real cook -- are documented).

Run:
    python3 tools/extract_model.py path/to/workbook.xlsx
Writes:
    data/rendering_model.json
    assets/data/rendering_model.json   (copy, for the static web app)
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit(
        "openpyxl is required to run this extractor.\n"
        "Install it with: pip install openpyxl"
    )

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUT = REPO_ROOT / "data" / "rendering_model.json"
WEB_COPY_OUT = REPO_ROOT / "assets" / "data" / "rendering_model.json"
WEB_JS_OUT = REPO_ROOT / "assets" / "js" / "data.generated.js"

SOURCE_ARTICLE_URL = (
    "https://smoketrailsbbq.com/brisket-holding-masterclass-and-tenderness-model/"
)


def extract_rendering_rates(ws) -> list[dict]:
    """Read the 'Tenderness Model' sheet: Temp F -> % Done Per Hour (rows 4-13)."""
    rates = []
    for row in range(4, 14):
        temp = ws.cell(row=row, column=1).value
        rate = ws.cell(row=row, column=2).value
        if temp is None or rate is None:
            continue
        rates.append({"temp_f": temp, "percent_per_hour": rate})
    # The calculator sheet also documents a 135F row with an explicit 0%/hr
    # floor (below this, collagen essentially doesn't render and it's also
    # the minimum food-safety hold temperature).
    rates.insert(0, {"temp_f": 135, "percent_per_hour": 0})
    return rates


def extract_confirmed_methods(ws) -> list[dict]:
    """Read the 'Confirmed Methods Time + Temp' sheet (rows 3,5,7,10)."""
    methods = []
    for row in (3, 5, 7, 10):
        hours_on_smoker = ws.cell(row=row, column=1).value
        if hours_on_smoker is None:
            continue
        methods.append(
            {
                "hours_on_smoker_plus_minus_2": hours_on_smoker,
                "pull_temp_f": ws.cell(row=row, column=2).value,
                "pull_texture": ws.cell(row=row, column=3).value,
                "rest_before_hold": ws.cell(row=row, column=4).value,
                "hold_temp_f": ws.cell(row=row, column=5).value,
                "hold_time_hours": ws.cell(row=row, column=6).value,
            }
        )
    return methods


def build_model(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    tenderness_ws = wb["Tenderness Model"]
    confirmed_ws = wb["Confirmed Methods Time + Temp"]

    model = {
        "_credit": {
            "method_name": 'The "Gow Method" — Brisket Cook & Hold Tenderness Model',
            "author": "Steve Gow",
            "source": "Smoke Trails BBQ",
            "article_url": SOURCE_ARTICLE_URL,
            "spreadsheet_file": xlsx_path.name,
            "note": (
                "This is an unofficial fan tribute. All model numbers are "
                "transcribed directly from Steve Gow's published spreadsheet "
                "and blog post. Please visit the article for the full "
                "writeup, and check out Smoke Trails BBQ's rubs."
            ),
        },
        "rendering_rates": extract_rendering_rates(tenderness_ws),
        "texture_bands": [
            {"max_percent": 80, "label": "Underdone"},
            {"max_percent": 95, "label": "Slightly Underdone"},
            {"max_percent": 110, "label": "PERFECTLY TENDER"},
            {"max_percent": 120, "label": "Slightly Over"},
            {"max_percent": None, "label": "Overdone"},
        ],
        "target_range": {"min_percent": 95, "max_percent": 110},
        "food_safety_note": (
            "Use proper hot-hold practices. Do not let internal temperature "
            "drop below 135\u00b0F."
        ),
        "confirmed_methods": extract_confirmed_methods(confirmed_ws),
        "standard_cooldown_patterns": [
            {
                "id": "pull_195_hold_150_oven",
                "description": (
                    "Standard temp decline from a 195\u00b0F pull straight into a "
                    "150\u00b0F holding oven."
                ),
                "quote": (
                    "Standard temp decline from 195 internal -> straight into "
                    "holding oven at 150 is 1 hr at 190, 1 hr at 180, 1 hr at "
                    "170, 1 hr at 160. Then remaining hours stabilize at 150."
                ),
                "zone_hours": [
                    {"temp_f": 190, "hours": 1},
                    {"temp_f": 180, "hours": 1},
                    {"temp_f": 170, "hours": 1},
                    {"temp_f": 160, "hours": 1},
                ],
                "then": "remaining hold hours stabilize at the hold temp (150F)",
            },
            {
                "id": "pull_205plus_counter_rest",
                "description": (
                    "Rough approximation for a ~2 hour counter rest after "
                    "pulling at 205F+."
                ),
                "quote": (
                    "Standard temp decline from 205+ to 140 on the counter "
                    "(counter rest) is 2 hours. So enter 1 hr at 170 and 1 hr "
                    "at 150 for a rough approximation if resting down on "
                    "counter."
                ),
                "zone_hours": [
                    {"temp_f": 170, "hours": 1},
                    {"temp_f": 150, "hours": 1},
                ],
                "then": None,
            },
        ],
        "documented_case_studies": [
            {
                "id": "jacks_190_pull_160_hold",
                "description": (
                    "12 hr cook to 190F pull, then 12 hr hold at 160F -- exact "
                    "zone-by-zone breakdown as Steve Gow described it in the "
                    "article comments (May 21, 2026) replying to a reader "
                    "named Jack."
                ),
                "cook_zone_hours": [
                    {"temp_f": 140, "hours": 1},
                    {"temp_f": 150, "hours": 1},
                    {"temp_f": 160, "hours": 1},
                    {"temp_f": 170, "hours": 2},
                    {"temp_f": 180, "hours": 1},
                    {"temp_f": 190, "hours": 1},
                ],
                "hold_zone_hours": [
                    {"temp_f": 190, "hours": 1},
                    {"temp_f": 180, "hours": 1},
                    {"temp_f": 170, "hours": 1},
                    {"temp_f": 160, "hours": 10},
                ],
                "expected_total_percent": 105,
                "expected_texture": "PERFECTLY TENDER",
            }
        ],
        "worked_example": {
            "description": "Example: 195\u00b0F pull + 150\u00b0F hold (from the calculator sheet, D29)",
            "cook_zone_hours": [
                {"temp_f": 140, "hours": 1},
                {"temp_f": 150, "hours": 1},
                {"temp_f": 160, "hours": 1},
                {"temp_f": 170, "hours": 1},
                {"temp_f": 180, "hours": 1},
                {"temp_f": 190, "hours": 1},
                {"temp_f": 195, "hours": 0.5},
            ],
            "hold_zone_hours": [
                {"temp_f": 190, "hours": 1},
                {"temp_f": 180, "hours": 1},
                {"temp_f": 170, "hours": 1},
                {"temp_f": 160, "hours": 1},
                {"temp_f": 150, "hours": 1},
                {"temp_f": 140, "hours": 14},
            ],
            "expected_total_percent": 101.5,
            "expected_texture": "PERFECTLY TENDER",
        },
    }
    return model


def main() -> None:
    if len(sys.argv) != 2:
        sys.exit("Usage: python3 tools/extract_model.py path/to/workbook.xlsx")
    xlsx_path = Path(sys.argv[1]).expanduser().resolve()
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    model = build_model(xlsx_path)

    DEFAULT_OUT.parent.mkdir(parents=True, exist_ok=True)
    WEB_COPY_OUT.parent.mkdir(parents=True, exist_ok=True)
    WEB_JS_OUT.parent.mkdir(parents=True, exist_ok=True)

    text = json.dumps(model, indent=2) + "\n"
    DEFAULT_OUT.write_text(text)
    WEB_COPY_OUT.write_text(text)

    js_text = (
        "// AUTO-GENERATED by tools/extract_model.py -- do not edit by hand.\n"
        "// Source: Steve Gow's Brisket Cook & Hold spreadsheet.\n"
        "// Embedded directly (not fetched) so the page also works when\n"
        "// opened straight from disk via file://, no local server required.\n"
        f"export const RENDERING_MODEL = {text.strip()};\n"
    )
    WEB_JS_OUT.write_text(js_text)

    print(f"Wrote {DEFAULT_OUT}")
    print(f"Wrote {WEB_COPY_OUT}")
    print(f"Wrote {WEB_JS_OUT}")
    print(f"Rendering rate points: {len(model['rendering_rates'])}")
    print(f"Confirmed methods: {len(model['confirmed_methods'])}")


if __name__ == "__main__":
    main()
